import csv
import io
from io import StringIO
from abc import ABC, abstractmethod
from openpyxl import load_workbook
from odf.opendocument import load as odf_load
from odf.table import Table, TableRow, TableCell
import xlrd
import dbf


class SpreadsheetExtractor(ABC):
    """
    Abstract base class for spreadsheet extractors.
    """

    @property
    def file_type(self) -> str:
        return "Spreadsheet"

    @abstractmethod
    def extract(self, buffer: bytes) -> str:
        """
        Abstract method to extract text from a spreadsheet.

        Args:
            buffer (bytes): The file content.

        Returns:
            str: Extracted spreadsheet content as text.
        """
        pass


class CSVExtractor(SpreadsheetExtractor):
    """
    Handles text extraction from CSV and TSV files.
    """

    def __init__(self, delimiter=","):
        """
        Initializes the extractor with the specified delimiter.

        Args:
            delimiter (str): Delimiter used in the file (e.g., ',' for CSV, '\t' for TSV).
        """
        self.delimiter = delimiter

    def extract(self, buffer: bytes) -> str:
        try:
            # Decode the file as text
            text = buffer.decode("utf-8", errors="replace")
            # Parse the CSV/TSV content using the specified delimiter
            reader = csv.reader(StringIO(text), delimiter=self.delimiter)
            return "\n".join([",".join(row) for row in reader])
        except Exception as e:
            raise ValueError(f"Failed to extract text from CSV/TSV: {e}")


class XLSXExtractor(SpreadsheetExtractor):
    """
    Handles text extraction from XLSX files.
    """

    def extract(self, buffer: bytes) -> str:
        try:
            # Load the workbook from the buffer
            workbook = load_workbook(io.BytesIO(buffer), read_only=True)
            text = []
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    text.append(",".join(map(str, row)))
            return "\n".join(text)
        except Exception as e:
            raise ValueError(f"Failed to extract text from XLSX: {e}")


class ODSExtractor(SpreadsheetExtractor):
    """
    Handles text extraction from ODS files.
    """

    def extract(self, buffer: bytes) -> str:
        try:
            # Load the ODS file
            ods_doc = odf_load(io.BytesIO(buffer))
            text = []
            for table in ods_doc.getElementsByType(Table):
                for row in table.getElementsByType(TableRow):
                    cells = [
                        cell.textContent
                        for cell in row.getElementsByType(TableCell)
                        if cell.textContent
                    ]
                    text.append(",".join(cells))
            return "\n".join(text)
        except Exception as e:
            raise ValueError(f"Failed to extract text from ODS: {e}")


class XLSExtractor(SpreadsheetExtractor):
    """
    Handles text extraction from legacy XLS files.
    """

    def extract(self, buffer: bytes) -> str:
        try:
            # Open XLS workbook
            workbook = xlrd.open_workbook(file_contents=buffer)
            text = []
            for sheet in workbook.sheets():
                for row_idx in range(sheet.nrows):
                    row = [str(cell) for cell in sheet.row(row_idx)]
                    text.append(",".join(row))
            return "\n".join(text)
        except Exception as e:
            raise ValueError(f"Failed to extract text from XLS: {e}")


class DBFExtractor(SpreadsheetExtractor):
    """
    Handles text extraction from DBF files.
    """

    def extract(self, buffer: bytes) -> str:
        try:
            # Open DBF table
            dbf_table = dbf.Table(io.BytesIO(buffer))
            text = []
            for record in dbf_table:
                text.append(",".join(map(str, record)))
            return "\n".join(text)
        except Exception as e:
            raise ValueError(f"Failed to extract text from DBF: {e}")
